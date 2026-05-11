#!/usr/bin/env python3
"""
Daily Shipping Tracking Report
- Scrapes every tracking link across all tabs in the Excel file
- Writes "Current Status" → column X, "Days in Transit" → column Y
- Skips rows already marked Delivered (status won't change)
- Runs 20 browser tabs in parallel for speed
- Saves progress every 100 rows so data isn't lost if interrupted
- Always uses today's date for in-transit calculations
- Auto-picks the newest .xlsx in the folder (or pass a path as an argument)

Tab filtering: only processes tabs matching these patterns (any week number):
  Failed_Delivery_W#, Delayed_Delivery_W#, Out_for_delivery_W#
  Tabs with no tracking data are automatically skipped.
"""

import asyncio
import re
import sys
import glob
import os
import openpyxl
from datetime import datetime, date
from playwright.async_api import async_playwright

FOLDER      = "/Users/nevafrazier/Desktop/Shipping Agent"
TODAY       = date.today()
COL_LINK    = 12   # L
COL_STATUS  = 24   # X
COL_DAYS    = 25   # Y
CONCURRENCY = 20   # parallel browser tabs
SAVE_EVERY  = 100  # save workbook after every N completions

TARGET_TABS = re.compile(
    r'^(Failed_Delivery|Delayed_Delivery|Out_for_delivery)_W\d+$',
    re.IGNORECASE,
)

def should_process_sheet(name):
    """Return True if this sheet should be scraped."""
    return bool(TARGET_TABS.match(name))

# ── helpers ──────────────────────────────────────────────────────────────────

def safe_save(wb, path):
    """Save workbook atomically — writes to a temp file first, then renames.
    This means if the process is killed mid-save, the original file stays intact."""
    import shutil
    tmp = path + '.tmp'
    wb.save(tmp)
    shutil.move(tmp, path)


def find_excel():
    if len(sys.argv) > 1:
        return sys.argv[1]
    files = [f for f in glob.glob(f"{FOLDER}/*.xlsx")
             if not os.path.basename(f).startswith('~$')]
    if not files:
        raise FileNotFoundError(f"No .xlsx files found in {FOLDER}")
    return max(files, key=lambda f: os.path.getmtime(f))


def parse_events(text):
    """Parse tracking timeline rows from page inner_text."""
    events = []
    lines  = text.split('\n')

    # Find the "Date/Time" header line
    start = None
    for i, line in enumerate(lines):
        if 'Date/Time' in line:
            start = i + 1
            break
    if start is None:
        return events

    stop_words = {'Hide Tracking Details', 'Contact', 'Terms', 'Privacy'}
    current_date = None
    i = start
    while i < len(lines):
        line = lines[i].strip()
        i += 1
        if not line:
            continue
        if line in stop_words:
            break
        # Date line: DD/MM/YYYY
        if re.match(r'^\d{2}/\d{2}/\d{4}$', line):
            current_date = line
            continue
        # Event line: "HH:MM AM/PM\tActivity\tLocation"
        if current_date and '\t' in line:
            parts = [p.strip() for p in line.split('\t')]
            if len(parts) >= 2:
                events.append({
                    'date':     current_date,
                    'time':     parts[0],
                    'activity': parts[1],
                    'location': parts[2] if len(parts) > 2 else '',
                })
    return events


def parse_date(s):
    try:
        return datetime.strptime(s, '%d/%m/%Y').date()
    except (ValueError, TypeError):
        return None


# Valid top-level status labels the page can show
VALID_STATUSES = {
    'Delivered', 'In Transit', 'Awaiting Collection',
    'Returned', 'Delivery Attempted', 'Out For Delivery',
}

def extract_page_status(text):
    """
    Pull the overall status badge from the page.
    It appears in ALL CAPS immediately after the tracking number line, e.g.:
        Tracking number
        889843324603
        IN TRANSIT          <- this is what we want
        Order number
    Returns title-cased string like 'In Transit', or None if not found.
    """
    lines = [l.strip() for l in text.split('\n') if l.strip()]
    for i, line in enumerate(lines):
        if line == 'Tracking number' and i + 2 < len(lines):
            candidate = lines[i + 2]
            if candidate.isupper() and len(candidate) > 1:
                return candidate.title()
    return None


def calc_days(status, pickup_date, delivered_date):
    if pickup_date is None:
        return 0
    if status.lower().strip() == 'delivered' and delivered_date:
        return (delivered_date - pickup_date).days
    return (TODAY - pickup_date).days


# ── playwright scraper ────────────────────────────────────────────────────────

async def scrape(page, url):
    """Return (current_status, pickup_date, delivered_date) or (None, None, None)."""
    await page.goto(url, wait_until='networkidle', timeout=30000)

    # Expand full history to see all events including Picked Up
    try:
        btn = page.locator('text=Show More').first
        if await btn.is_visible(timeout=2000):
            await btn.click()
            await page.wait_for_timeout(1500)
    except Exception:
        pass

    text   = await page.inner_text('body')

    # ── Dates: parse the event timeline for Picked Up and Delivered timestamps
    events = parse_events(text)

    pickup_date = next(
        (parse_date(e['date']) for e in events
         if e['activity'].lower().strip().startswith('picked up')),
        None
    )
    delivered_date = next(
        (parse_date(e['date']) for e in events
         if e['activity'].lower().strip().startswith('delivered')),
        None
    )

    # ── Overall status: read the badge that sits right after the tracking number.
    # The badge is the authoritative single label for the package state.
    # E.g. "IN TRANSIT" → "In Transit",  "AWAITING COLLECTION" → "Awaiting Collection"
    current_status = extract_page_status(text)

    if current_status is None:
        # Fallback: derive from events if badge couldn't be read
        if delivered_date is not None:
            current_status = 'Delivered'
        elif any('return' in e['activity'].lower() for e in events):
            current_status = 'Returned'
        elif events:
            current_status = events[0]['activity']   # raw event text as last resort

    # Normalize return-related labels to "Returned"
    if current_status and current_status.lower().strip() in ('returns', 'return'):
        current_status = 'Returned'

    return current_status, pickup_date, delivered_date


# ── main ──────────────────────────────────────────────────────────────────────

async def main():
    excel_path = find_excel()
    print(f"Today : {TODAY}")
    print(f"File  : {excel_path}")
    print()

    wb = openpyxl.load_workbook(excel_path)

    # Set column headers on every sheet and collect rows to scrape
    work = []   # list of (sheet_name, row_number, url)
    skipped = 0

    mode_label = "official (Failed_Delivery_W#, Delayed_Delivery_W#, Out_for_delivery_W#)"
    print(f"Mode  : {mode_label}")
    print()

    for sheet_name in wb.sheetnames:
        if not should_process_sheet(sheet_name):
            continue

        ws = wb[sheet_name]

        # Always write headers in row 1
        ws.cell(1, COL_STATUS).value = "Current Status"
        ws.cell(1, COL_DAYS).value   = "Days in Transit"

        # Skip sheets with no tracking link column
        if ws.cell(1, COL_LINK).value is None:
            continue

        for row in range(2, ws.max_row + 1):
            url = ws.cell(row, COL_LINK).value
            if not url:
                continue
            existing = ws.cell(row, COL_STATUS).value
            # Only skip rows that are cleanly "Delivered" — that status never changes.
            # Re-scrape anything blank or set to a raw event description from a bad run.
            if existing and str(existing).strip() == 'Delivered':
                skipped += 1
                continue
            work.append((sheet_name, row, url))

    total = len(work)
    print(f"Rows to scrape : {total:,}")
    print(f"Already done   : {skipped:,} (Delivered — skipped)")
    print(f"Concurrency    : {CONCURRENCY} parallel tabs")
    print()

    if total == 0:
        wb.save(excel_path)
        print("Nothing new to scrape. File saved.")
        return

    # Shared state (all access is single-threaded in asyncio)
    completed = 0
    errors    = 0
    sem       = asyncio.Semaphore(CONCURRENCY)
    save_lock = asyncio.Lock()

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)

        async def process(sheet_name, row, url):
            nonlocal completed, errors
            async with sem:
                status, pickup, delivered = None, None, None
                for attempt in range(3):
                    page = await browser.new_page()
                    try:
                        status, pickup, delivered = await scrape(page, url)
                        if status is not None:
                            break  # success
                    except Exception:
                        pass
                    finally:
                        await page.close()
                    await asyncio.sleep(1)  # brief pause before retry

                if status is not None:
                    days = calc_days(status, pickup, delivered)
                    wb[sheet_name].cell(row, COL_STATUS).value = status
                    wb[sheet_name].cell(row, COL_DAYS).value   = days
                else:
                    errors += 1
                    wb[sheet_name].cell(row, COL_STATUS).value = 'Check Manually'
                    wb[sheet_name].cell(row, COL_DAYS).value   = ''

                completed += 1

                if completed % 10 == 0 or completed == total:
                    pct = completed / total * 100
                    print(f"  {completed:>6,}/{total:,}  ({pct:5.1f}%)  errors={errors}", flush=True)

                if completed % SAVE_EVERY == 0:
                    async with save_lock:
                        safe_save(wb, excel_path)
                        print(f"  [saved at {completed:,} rows]", flush=True)

        await asyncio.gather(*[process(s, r, u) for s, r, u in work])
        await browser.close()

    safe_save(wb, excel_path)
    print()
    print(f"Finished! {completed:,} rows processed, {errors:,} errors.")
    print(f"Saved : {excel_path}")


if __name__ == '__main__':
    asyncio.run(main())
